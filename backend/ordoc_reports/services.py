import os
import json
import csv
import logging
from datetime import datetime, timedelta
from django.utils import timezone
from django.db import models
from django.apps import apps
from django.core.files.storage import default_storage
from django.conf import settings
from django.core.mail import send_mail
import tempfile
import zipfile
import tarfile
from io import BytesIO, StringIO
import requests

# Importações opcionais para geração de relatórios
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class NotificationService:
    """
    Serviço utilitário para envio de notificações (e-mail e webhook)
    """

    logger = logging.getLogger(__name__)

    @staticmethod
    def send_email(subject, message, recipients):
        if not recipients:
            NotificationService.logger.info(
                "Nenhum destinatário informado para e-mail de notificação"
            )
            return

        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"),
                recipient_list=recipients,
                fail_silently=False,
            )
            NotificationService.logger.info(
                f"E-mail de notificação enviado para {recipients}"
            )
        except Exception as exc:
            NotificationService.logger.error(
                f"Falha ao enviar e-mail de notificação para {recipients}: {exc}"
            )
            raise

    @staticmethod
    def send_webhook(url, payload):
        try:
            response = requests.post(url, json=payload, timeout=10)
            response.raise_for_status()
            NotificationService.logger.info(
                f"Webhook de notificação enviado para {url} (status {response.status_code})"
            )
        except Exception as exc:
            NotificationService.logger.error(
                f"Falha ao enviar webhook de notificação para {url}: {exc}"
            )
            raise

    @staticmethod
    def notify(subject, message, emails=None, webhook_url=None, payload=None):
        """Envia notificações por e-mail e/ou webhook"""
        if emails:
            NotificationService.send_email(subject, message, emails)

        if webhook_url:
            payload = payload or {"subject": subject, "message": message}
            NotificationService.send_webhook(webhook_url, payload)


class ReportGenerationService:
    """
    Serviço para geração de relatórios
    Responsável por executar consultas e gerar arquivos de relatórios
    """
    
    @staticmethod
    def generate_preview(template, filters=None, organization=None, limit=10):
        """
        Gera uma prévia limitada do relatório
        """
        try:
            # Executa a consulta com limite
            data = ReportGenerationService._execute_query(
                template.query_config,
                filters or {},
                organization,
                limit=limit
            )
            
            return {
                'columns': data.get('columns', []),
                'rows': data.get('rows', [])[:limit],
                'total_rows': len(data.get('rows', [])),
                'preview': True
            }
        
        except Exception as e:
            raise Exception(f"Erro ao gerar prévia: {str(e)}")
    
    @staticmethod
    def generate_report(report):
        """
        Gera um relatório completo
        """
        start_time = timezone.now()
        
        try:
            # Atualiza status
            report.status = 'generating'
            report.save()
            
            # Executa a consulta
            data = ReportGenerationService._execute_query(
                report.template.query_config,
                report.filters_applied,
                report.organization
            )
            
            # Aplica configurações de exibição
            formatted_data = ReportGenerationService._apply_display_config(
                data,
                report.template.display_config
            )
            
            # Gera arquivo no formato solicitado
            file_path = ReportGenerationService._generate_file(
                formatted_data,
                report.format,
                report.title,
                report.template.export_config
            )
            
            # Calcula tempo de geração
            generation_time = timezone.now() - start_time
            
            # Atualiza relatório
            report.status = 'completed'
            report.data = formatted_data
            report.file_path = file_path
            report.file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            report.generation_time = generation_time
            report.metadata = {
                'rows_count': len(formatted_data.get('rows', [])),
                'columns_count': len(formatted_data.get('columns', [])),
                'generated_at': timezone.now().isoformat(),
                'generation_time_seconds': generation_time.total_seconds()
            }
            report.save()
            
            # Registra métrica de tempo de geração
            ReportGenerationService._record_metric(
                report,
                'generation_time',
                generation_time.total_seconds(),
                'seconds'
            )
            
            return report
        
        except Exception as e:
            # Atualiza status de erro
            report.status = 'failed'
            report.error_message = str(e)
            report.generation_time = timezone.now() - start_time
            report.save()
            
            # Registra métrica de erro
            ReportGenerationService._record_metric(
                report,
                'error_rate',
                1,
                'count'
            )
            
            raise e
    
    @staticmethod
    def _execute_query(query_config, filters, organization, limit=None):
        """
        Executa a consulta baseada na configuração
        """
        try:
            # Obtém o modelo
            model_name = query_config.get('model')
            if not model_name:
                raise ValueError("Modelo não especificado na configuração da consulta")
            
            # Tenta obter o modelo
            try:
                app_label, model_class = model_name.split('.')
                Model = apps.get_model(app_label, model_class)
            except (ValueError, LookupError):
                raise ValueError(f"Modelo '{model_name}' não encontrado")
            
            # Constrói a consulta
            queryset = Model.objects.all()
            
            # Aplica filtro de organização se disponível
            if organization and hasattr(Model, 'organization'):
                queryset = queryset.filter(organization=organization)
            
            # Aplica filtros personalizados
            if filters:
                for field, value in filters.items():
                    if hasattr(Model, field) and value:
                        # Suporte a filtros básicos
                        if field.endswith('__gte'):
                            queryset = queryset.filter(**{field: value})
                        elif field.endswith('__lte'):
                            queryset = queryset.filter(**{field: value})
                        elif field.endswith('__contains'):
                            queryset = queryset.filter(**{field: value})
                        else:
                            queryset = queryset.filter(**{field: value})
            
            # Aplica filtros da configuração da consulta
            query_filters = query_config.get('filters', {})
            if query_filters:
                queryset = queryset.filter(**query_filters)
            
            # Aplica ordenação
            ordering = query_config.get('ordering', [])
            if ordering:
                queryset = queryset.order_by(*ordering)
            
            # Aplica limite se especificado
            if limit:
                queryset = queryset[:limit]
            
            # Obtém campos solicitados
            fields = query_config.get('fields', [])
            if not fields:
                raise ValueError("Campos não especificados na configuração da consulta")
            
            # Executa a consulta
            if query_config.get('use_values', True):
                # Usa values() para melhor performance
                data = list(queryset.values(*fields))
            else:
                # Usa objetos completos
                data = []
                for obj in queryset:
                    row = {}
                    for field in fields:
                        if hasattr(obj, field):
                            value = getattr(obj, field)
                            # Converte valores especiais
                            if hasattr(value, 'isoformat'):  # datetime
                                value = value.isoformat()
                            elif hasattr(value, '__str__'):
                                value = str(value)
                            row[field] = value
                    data.append(row)
            
            # Formata resultado
            columns = [{'name': field, 'label': field.replace('_', ' ').title()} for field in fields]
            rows = data
            
            return {
                'columns': columns,
                'rows': rows,
                'total_rows': len(rows)
            }
        
        except Exception as e:
            raise Exception(f"Erro ao executar consulta: {str(e)}")
    
    @staticmethod
    def _apply_display_config(data, display_config):
        """
        Aplica configurações de exibição aos dados
        """
        if not display_config:
            return data
        
        # Aplica formatação de colunas
        column_formats = display_config.get('column_formats', {})
        if column_formats:
            for row in data['rows']:
                for field, format_config in column_formats.items():
                    if field in row:
                        row[field] = ReportGenerationService._format_value(
                            row[field], 
                            format_config
                        )
        
        # Aplica labels personalizados
        column_labels = display_config.get('column_labels', {})
        if column_labels:
            for column in data['columns']:
                if column['name'] in column_labels:
                    column['label'] = column_labels[column['name']]
        
        return data
    
    @staticmethod
    def _format_value(value, format_config):
        """
        Formata um valor baseado na configuração
        """
        if value is None:
            return format_config.get('null_value', '')
        
        format_type = format_config.get('type', 'string')
        
        if format_type == 'currency':
            try:
                return f"R$ {float(value):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except (ValueError, TypeError):
                return str(value)
        
        elif format_type == 'percentage':
            try:
                return f"{float(value):.2f}%"
            except (ValueError, TypeError):
                return str(value)
        
        elif format_type == 'date':
            try:
                if isinstance(value, str):
                    from dateutil import parser
                    value = parser.parse(value)
                return value.strftime(format_config.get('format', '%d/%m/%Y'))
            except:
                return str(value)
        
        elif format_type == 'datetime':
            try:
                if isinstance(value, str):
                    from dateutil import parser
                    value = parser.parse(value)
                return value.strftime(format_config.get('format', '%d/%m/%Y %H:%M'))
            except:
                return str(value)
        
        return str(value)
    
    @staticmethod
    def _generate_file(data, format, title, export_config):
        """
        Gera arquivo no formato especificado
        """
        # Cria diretório de relatórios se não existir
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Nome do arquivo
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{title}_{timestamp}.{format}"
        file_path = os.path.join(reports_dir, filename)
        
        if format == 'csv':
            return ReportGenerationService._generate_csv(data, file_path)
        elif format == 'excel':
            return ReportGenerationService._generate_excel(data, file_path, title, export_config)
        elif format == 'pdf':
            return ReportGenerationService._generate_pdf(data, file_path, title, export_config)
        elif format == 'json':
            return ReportGenerationService._generate_json(data, file_path)
        elif format == 'html':
            return ReportGenerationService._generate_html(data, file_path, title, export_config)
        else:
            raise ValueError(f"Formato '{format}' não suportado")
    
    @staticmethod
    def _generate_csv(data, file_path):
        """Gera arquivo CSV"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            if not data['rows']:
                return file_path
            
            fieldnames = [col['name'] for col in data['columns']]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            # Escreve cabeçalho
            header = {col['name']: col['label'] for col in data['columns']}
            writer.writerow(header)
            
            # Escreve dados
            writer.writerows(data['rows'])
        
        return file_path
    
    @staticmethod
    def _generate_excel(data, file_path, title, export_config):
        """Gera arquivo Excel (ou CSV se Excel não disponível)"""
        if not EXCEL_AVAILABLE:
            # Fallback para CSV se Excel não estiver disponível
            csv_path = file_path.replace('.excel', '.csv')
            return ReportGenerationService._generate_csv(data, csv_path)
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = title[:31]  # Limite do Excel
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escreve cabeçalho
        for col_idx, column in enumerate(data['columns'], 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escreve dados
        for row_idx, row_data in enumerate(data['rows'], 2):
            for col_idx, column in enumerate(data['columns'], 1):
                value = row_data.get(column['name'], '')
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajusta largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(file_path)
        return file_path
    
    @staticmethod
    def _generate_pdf(data, file_path, title, export_config):
        """Gera arquivo PDF (ou HTML se PDF não disponível)"""
        if not PDF_AVAILABLE:
            # Fallback para HTML se PDF não estiver disponível
            html_path = file_path.replace('.pdf', '.html')
            return ReportGenerationService._generate_html(data, html_path, title, export_config)
        
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=30
        )
        elements.append(Paragraph(title, title_style))
        
        # Data de geração
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,  # Center
            spaceAfter=20
        )
        elements.append(Paragraph(f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}", date_style))
        
        if data['rows']:
            # Prepara dados da tabela
            table_data = []
            
            # Cabeçalho
            headers = [col['label'] for col in data['columns']]
            table_data.append(headers)
            
            # Dados
            for row in data['rows']:
                table_row = [str(row.get(col['name'], '')) for col in data['columns']]
                table_data.append(table_row)
            
            # Cria tabela
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        else:
            elements.append(Paragraph("Nenhum dado encontrado", styles['Normal']))
        
        doc.build(elements)
        return file_path
    
    @staticmethod
    def _generate_json(data, file_path):
        """Gera arquivo JSON"""
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(data, jsonfile, ensure_ascii=False, indent=2, default=str)
        
        return file_path
    
    @staticmethod
    def _generate_html(data, file_path, title, export_config):
        """Gera arquivo HTML"""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; text-align: center; }}
                .meta {{ text-align: center; color: #666; margin-bottom: 30px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <div class="meta">Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')}</div>
        """
        
        if data['rows']:
            html_content += "<table>"
            
            # Cabeçalho
            html_content += "<thead><tr>"
            for column in data['columns']:
                html_content += f"<th>{column['label']}</th>"
            html_content += "</tr></thead>"
            
            # Dados
            html_content += "<tbody>"
            for row in data['rows']:
                html_content += "<tr>"
                for column in data['columns']:
                    value = row.get(column['name'], '')
                    html_content += f"<td>{value}</td>"
                html_content += "</tr>"
            html_content += "</tbody>"
            
            html_content += "</table>"
        else:
            html_content += "<p>Nenhum dado encontrado</p>"
        
        html_content += """
        </body>
        </html>
        """
        
        with open(file_path, 'w', encoding='utf-8') as htmlfile:
            htmlfile.write(html_content)
        
        return file_path
    
    @staticmethod
    def _record_metric(report, metric_type, value, unit):
        """Registra uma métrica do relatório"""
        from .models import ReportMetric
        
        ReportMetric.objects.create(
            metric_type=metric_type,
            metric_name=f"{report.template.name} - {metric_type}",
            metric_value=value,
            metric_unit=unit,
            period_start=timezone.now(),
            period_end=timezone.now(),
            report=report,
            report_template=report.template,
            organization=report.organization
        )


class ReportExportService:
    """
    Serviço para exportação em lote de relatórios
    """
    
    @staticmethod
    def export_reports(report_ids, export_format='zip', include_metadata=True):
        """
        Exporta múltiplos relatórios em um arquivo compactado
        """
        from .models import Report
        
        reports = Report.objects.filter(id__in=report_ids, status='completed')
        
        if not reports.exists():
            raise ValueError("Nenhum relatório válido encontrado")
        
        # Cria arquivo temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{export_format}') as temp_file:
            if export_format == 'zip':
                return ReportExportService._create_zip(reports, temp_file.name, include_metadata)
            elif export_format == 'tar':
                return ReportExportService._create_tar(reports, temp_file.name, include_metadata)
            else:
                raise ValueError(f"Formato de exportação '{export_format}' não suportado")
    
    @staticmethod
    def _create_zip(reports, output_path, include_metadata):
        """Cria arquivo ZIP com os relatórios"""
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for report in reports:
                if report.file_path and os.path.exists(report.file_path):
                    # Adiciona arquivo do relatório
                    arcname = f"{report.title}.{report.format}"
                    zipf.write(report.file_path, arcname)
                    
                    # Adiciona metadados se solicitado
                    if include_metadata:
                        metadata = {
                            'id': str(report.id),
                            'title': report.title,
                            'description': report.description,
                            'template': report.template.name,
                            'generated_by': report.generated_by.name,
                            'created_at': report.created_at.isoformat(),
                            'filters_applied': report.filters_applied,
                            'parameters': report.parameters,
                            'metadata': report.metadata
                        }
                        
                        metadata_json = json.dumps(metadata, ensure_ascii=False, indent=2, default=str)
                        zipf.writestr(f"{report.title}_metadata.json", metadata_json)
        
        return output_path
    
    @staticmethod
    def _create_tar(reports, output_path, include_metadata):
        """Cria arquivo TAR com os relatórios"""
        with tarfile.open(output_path, 'w:gz') as tarf:
            for report in reports:
                if report.file_path and os.path.exists(report.file_path):
                    # Adiciona arquivo do relatório
                    arcname = f"{report.title}.{report.format}"
                    tarf.add(report.file_path, arcname=arcname)
                    
                    # Adiciona metadados se solicitado
                    if include_metadata:
                        metadata = {
                            'id': str(report.id),
                            'title': report.title,
                            'description': report.description,
                            'template': report.template.name,
                            'generated_by': report.generated_by.name,
                            'created_at': report.created_at.isoformat(),
                            'filters_applied': report.filters_applied,
                            'parameters': report.parameters,
                            'metadata': report.metadata
                        }
                        
                        metadata_json = json.dumps(metadata, ensure_ascii=False, indent=2, default=str)
                        
                        # Cria arquivo temporário para metadados
                        with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json') as temp_meta:
                            temp_meta.write(metadata_json)
                            temp_meta.flush()
                            tarf.add(temp_meta.name, arcname=f"{report.title}_metadata.json")
                            os.unlink(temp_meta.name)
        
        return output_path


class ReportScheduleService:
    """
    Serviço para gerenciamento de agendamentos de relatórios
    """

    logger = logging.getLogger(__name__)
    
    @staticmethod
    def process_scheduled_reports():
        """
        Processa relatórios agendados que devem ser executados
        """
        from .models import ReportSchedule, Report
        
        now = timezone.now()
        
        # Busca agendamentos ativos que devem ser executados
        schedules = ReportSchedule.objects.filter(
            status='active',
            next_run__lte=now
        )
        
        for schedule in schedules:
            try:
                # Cria relatório baseado no agendamento
                report = Report.objects.create(
                    title=f"{schedule.name} - {now.strftime('%Y-%m-%d %H:%M')}",
                    description=f"Relatório gerado automaticamente pelo agendamento: {schedule.name}",
                    format=schedule.default_format,
                    filters_applied=schedule.default_filters,
                    parameters={},
                    template=schedule.template,
                    organization=schedule.organization,
                    generated_by=schedule.created_by
                )
                
                # Define expiração
                report.set_default_expiry()
                report.save()
                
                # Inicia geração assíncrona
                from .tasks import generate_report_task
                generate_report_task.delay(report.id)
                
                # Atualiza agendamento
                schedule.last_run = now
                schedule.calculate_next_run()
                
                print(f"Relatório agendado criado: {report.id} para agendamento {schedule.name}")
            
            except Exception as e:
                print(f"Erro ao processar agendamento {schedule.name}: {str(e)}")
                
                # Opcional: enviar notificação de erro
                if schedule.notify_on_error:
                    ReportScheduleService._send_error_notification(schedule, str(e))
    
    @staticmethod
    def _send_error_notification(schedule, error_message):
        """
        Envia notificação de erro no agendamento
        """
        recipients = schedule.notification_emails or []

        if not recipients:
            ReportScheduleService.logger.info(
                f"Nenhum destinatário configurado para o agendamento {schedule.name}"
            )
            return

        subject = f"Erro no agendamento de relatório: {schedule.name}"
        message = (
            "Ocorreu um erro ao executar o agendamento:\n\n" f"{error_message}"
        )

        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"),
                recipient_list=recipients,
                fail_silently=False,
            )
            ReportScheduleService.logger.info(
                f"E-mail de erro enviado para {recipients}"
            )
        except Exception as exc:
            ReportScheduleService.logger.error(
                f"Falha ao enviar e-mail de erro para {recipients}: {exc}"
            )
            raise


class ReportCleanupService:
    """
    Serviço para limpeza de relatórios expirados
    """
    
    @staticmethod
    def cleanup_expired_reports():
        """
        Remove arquivos de relatórios expirados
        """
        from .models import Report
        
        now = timezone.now()
        
        # Busca relatórios expirados
        expired_reports = Report.objects.filter(
            expires_at__lt=now,
            file_path__isnull=False
        ).exclude(file_path='')
        
        cleaned_count = 0
        
        for report in expired_reports:
            try:
                if report.file_path and os.path.exists(report.file_path):
                    os.remove(report.file_path)
                    
                # Limpa dados do relatório
                report.file_path = None
                report.file_size = None
                report.data = {}
                report.save()
                
                cleaned_count += 1
                
            except Exception as e:
                print(f"Erro ao limpar relatório {report.id}: {str(e)}")
        
        print(f"Limpeza concluída: {cleaned_count} relatórios limpos")
        return cleaned_count
    
    @staticmethod
    def cleanup_old_reports(days=90):
        """
        Remove relatórios muito antigos (independente da expiração)
        """
        from .models import Report
        
        cutoff_date = timezone.now() - timedelta(days=days)
        
        old_reports = Report.objects.filter(
            created_at__lt=cutoff_date,
            file_path__isnull=False
        ).exclude(file_path='')
        
        cleaned_count = 0
        
        for report in old_reports:
            try:
                if report.file_path and os.path.exists(report.file_path):
                    os.remove(report.file_path)
                
                # Remove o relatório completamente
                report.delete()
                cleaned_count += 1
                
            except Exception as e:
                print(f"Erro ao remover relatório antigo {report.id}: {str(e)}")
        
        print(f"Limpeza de relatórios antigos concluída: {cleaned_count} relatórios removidos")
        return cleaned_count
